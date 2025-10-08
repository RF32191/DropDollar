#!/usr/bin/env python3
"""
macOS-Compatible Crypto Auto Broker
Automated cryptocurrency analysis and prediction system with email alerts
"""

import os
import json
import time
import math
import csv
import requests
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from pycoingecko import CoinGeckoAPI
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import platform
import logging

# =======================
# LOGGING SETUP
# =======================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('crypto_broker.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =======================
# CONFIG
# =======================
# Get user's desktop path for macOS
if platform.system() == "Darwin":  # macOS
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
else:
    desktop_path = os.path.expanduser("~/Desktop")

CONFIG = {
    "OUTPUT_FOLDER": os.path.join(desktop_path, "CryptoBroker"),
    "EMAILS": ["ryanfermoselle@outlook.com"],
    "EMAIL_SETTINGS": {
        "SMTP_SERVER": "smtp-mail.outlook.com",
        "SMTP_PORT": 587,
        "EMAIL_USER": "",  # Set this to your email
        "EMAIL_PASS": "",  # Set this to your app password
    },
    "USE_MINUTE_FOR_5M": True,
    "MAX_RETRIES": 5,  # Increased for better resilience
    "TIMEOUT": 15,     # Increased timeout
    "RETRY_DELAY": 2,  # Base delay between retries
    "ALIGN_TO_WALLCLOCK": True,
    "FALLBACK_YAHOO_PRICE": True,
    "LEARNING_RATE": 0.05,
    "EVAL_WINDOW_MIN": 4,
    "CONNECTION_RESILIENCE": True,  # Enhanced error handling
}

# Create output directory
os.makedirs(CONFIG["OUTPUT_FOLDER"], exist_ok=True)

# File paths
WEIGHTS_FILE = os.path.join(CONFIG["OUTPUT_FOLDER"], "weights.json")
PRED_STATE_FILE = os.path.join(CONFIG["OUTPUT_FOLDER"], "last_predictions.json")
LAST_STATE_FILE = os.path.join(CONFIG["OUTPUT_FOLDER"], "last_state.json")
EVAL_LOG = os.path.join(CONFIG["OUTPUT_FOLDER"], "eval_log.csv")
CONFIG_FILE = os.path.join(CONFIG["OUTPUT_FOLDER"], "email_config.json")

# Crypto mappings with additional coins for diversification
CRYPTO_TICKERS = {
    "BTC-USD": "bitcoin",
    "ETH-USD": "ethereum", 
    "SOL-USD": "solana",
    "ADA-USD": "cardano",
    "XRP-USD": "ripple",
    "MATIC-USD": "matic-network",
    "DOT-USD": "polkadot",
    "LINK-USD": "chainlink",
    "AVAX-USD": "avalanche-2",
    "ATOM-USD": "cosmos"
}

# Initialize CoinGecko API
cg = CoinGeckoAPI()

# =======================
# EMAIL CONFIG MANAGEMENT
# =======================
def load_email_config():
    """Load email configuration from file or create template"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config = json.load(f)
                CONFIG["EMAIL_SETTINGS"].update(config)
                return True
        except Exception as e:
            logger.error(f"Failed to load email config: {e}")
    
    # Create template config file
    template_config = {
        "EMAIL_USER": "your_email@outlook.com",
        "EMAIL_PASS": "your_app_password_here",
        "SMTP_SERVER": "smtp-mail.outlook.com",
        "SMTP_PORT": 587
    }
    
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(template_config, f, indent=2)
        logger.info(f"Created email config template at: {CONFIG_FILE}")
        logger.info("Please update the email configuration file with your credentials")
    except Exception as e:
        logger.error(f"Failed to create email config template: {e}")
    
    return False

# =======================
# DEFAULT WEIGHTS
# =======================
DEFAULT_PREDICT_WEIGHTS = {
    "m_r5": 0.45,
    "m_macd": 0.25,
    "m_rsi": 0.15,
    "m_r1": 0.15,
    "d_signal": 0.002,
    "max_5m": 0.004,
    "weekly_alpha": 0.6,
    "weekly_beta": 0.2,
    "weekly_cap": 0.15,
    "candle_w_min": 0.6,
    "candle_w_day": 0.4
}

def load_weights():
    """Load prediction weights from file"""
    if os.path.exists(WEIGHTS_FILE):
        try:
            with open(WEIGHTS_FILE, "r", encoding="utf-8") as f:
                w = json.load(f)
            for k, v in DEFAULT_PREDICT_WEIGHTS.items():
                if k not in w:
                    w[k] = v
            return w
        except Exception:
            pass
    return DEFAULT_PREDICT_WEIGHTS.copy()

PREDICT_WEIGHTS = load_weights()

# =======================
# ENHANCED HTTP & DATA FETCH
# =======================
def exponential_backoff_retry(fn, *args, **kwargs):
    """Enhanced retry mechanism with exponential backoff"""
    max_retries = CONFIG["MAX_RETRIES"]
    base_delay = CONFIG["RETRY_DELAY"]
    
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except requests.exceptions.ConnectionError as e:
            if attempt == max_retries - 1:
                logger.error(f"Connection failed after {max_retries} attempts: {e}")
                raise
            delay = base_delay * (2 ** attempt) + np.random.uniform(0, 1)
            logger.warning(f"Connection attempt {attempt + 1} failed, retrying in {delay:.2f}s")
            time.sleep(delay)
        except requests.exceptions.Timeout as e:
            if attempt == max_retries - 1:
                logger.error(f"Timeout after {max_retries} attempts: {e}")
                raise
            delay = base_delay * (2 ** attempt)
            logger.warning(f"Timeout on attempt {attempt + 1}, retrying in {delay:.2f}s")
            time.sleep(delay)
        except Exception as e:
            if attempt == max_retries - 1:
                logger.error(f"Request failed after {max_retries} attempts: {e}")
                raise
            delay = base_delay * (2 ** attempt)
            logger.warning(f"Request attempt {attempt + 1} failed: {e}, retrying in {delay:.2f}s")
            time.sleep(delay)

def fetch_current_prices_batched(coin_ids):
    """Fetch current prices with enhanced error handling"""
    def _batch_call():
        ids = ",".join(coin_ids)
        url = "https://api.coingecko.com/api/v3/simple/price"
        params = {
            "ids": ids, 
            "vs_currencies": "usd", 
            "include_24hr_vol": "true",
            "include_24hr_change": "true"
        }
        headers = {
            'User-Agent': 'CryptoBroker/1.0',
            'Accept': 'application/json'
        }
        
        response = requests.get(url, params=params, headers=headers, timeout=CONFIG["TIMEOUT"])
        response.raise_for_status()
        return response.json()
    
    try:
        data = exponential_backoff_retry(_batch_call)
        result = {}
        for cid in coin_ids:
            coin_data = data.get(cid, {})
            price = coin_data.get("usd")
            volume = coin_data.get("usd_24h_vol", 0)
            if price is not None:
                result[cid] = (float(price), float(volume))
            else:
                result[cid] = (None, None)
        return result
    except Exception as e:
        logger.error(f"Batch price fetch failed: {e}")
        # Fallback to individual requests
        result = {}
        for cid in coin_ids:
            try:
                def _single_call():
                    return cg.get_price(ids=cid, vs_currencies="usd", include_24hr_vol=True)
                
                res = exponential_backoff_retry(_single_call)
                coin_data = res.get(cid, {})
                price = float(coin_data.get("usd", 0))
                vol = float(coin_data.get("usd_24h_vol", 0))
                result[cid] = (price, vol)
                time.sleep(0.1)  # Rate limiting
            except Exception as single_e:
                logger.error(f"Failed to fetch {cid}: {single_e}")
                result[cid] = (None, None)
        return result

def fetch_historical_series(coin_id, days):
    """Fetch historical price series with enhanced error handling"""
    def _hist_call():
        # Remove interval parameter to avoid Enterprise plan requirement
        # CoinGecko automatically selects appropriate interval based on days
        return cg.get_coin_market_chart_by_id(
            id=coin_id, 
            vs_currency="usd", 
            days=days
        )
    
    try:
        hist = exponential_backoff_retry(_hist_call)
        prices = hist.get("prices", [])
        if not prices:
            logger.warning(f"No price data for {coin_id}")
            return pd.Series(dtype="float64")
        
        df = pd.DataFrame(prices, columns=["ts", "price"])
        df["ts"] = pd.to_datetime(df["ts"], unit="ms")
        df.set_index("ts", inplace=True)
        s = df["price"].astype(float).sort_index()
        
        # Remove any duplicate timestamps
        s = s[~s.index.duplicated(keep='last')]
        
        logger.debug(f"Fetched {len(s)} price points for {coin_id}")
        return s
    except Exception as e:
        logger.error(f"Failed to fetch historical data for {coin_id}: {e}")
        return pd.Series(dtype="float64")

def yfinance_last_price(symbol):
    """Fallback to Yahoo Finance for price data"""
    if not CONFIG["FALLBACK_YAHOO_PRICE"]:
        return None
    
    try:
        import yfinance as yf
        ticker = yf.Ticker(symbol)
        
        # Try fast_info first
        if hasattr(ticker, 'fast_info'):
            info = ticker.fast_info
            if hasattr(info, 'last_price'):
                return float(info.last_price)
        
        # Fallback to recent history
        df = ticker.history(period="1d", interval="1m")
        if df is not None and not df.empty:
            return float(df["Close"].dropna().iloc[-1])
        
        return None
    except Exception as e:
        logger.error(f"Yahoo Finance fallback failed for {symbol}: {e}")
        return None

# =======================
# TECHNICAL INDICATORS
# =======================
def calculate_rsi(prices, period=14):
    """Calculate RSI with improved handling"""
    try:
        p = prices.astype(float).dropna()
        if len(p) < period + 1:
            return pd.Series(dtype="float64", index=prices.index)
        
        delta = p.diff()
        gain = delta.clip(lower=0).rolling(window=period, min_periods=period).mean()
        loss = -delta.clip(upper=0).rolling(window=period, min_periods=period).mean()
        
        # Avoid division by zero
        rs = gain / loss.replace(0, np.nan)
        rsi = 100 - (100 / (1 + rs))
        
        return rsi.fillna(50)  # Neutral RSI for missing values
    except Exception as e:
        logger.error(f"RSI calculation failed: {e}")
        return pd.Series(dtype="float64", index=prices.index)

def calculate_macd(prices):
    """Calculate MACD histogram"""
    try:
        p = prices.astype(float).dropna()
        if len(p) < 26:
            return pd.Series(dtype="float64", index=prices.index)
        
        ema12 = p.ewm(span=12, adjust=False).mean()
        ema26 = p.ewm(span=26, adjust=False).mean()
        macd_line = ema12 - ema26
        signal_line = macd_line.ewm(span=9, adjust=False).mean()
        histogram = macd_line - signal_line
        
        return histogram
    except Exception as e:
        logger.error(f"MACD calculation failed: {e}")
        return pd.Series(dtype="float64", index=prices.index)

def calculate_stochastic_from_price(prices, k_period=14, d_period=3):
    """Calculate Stochastic oscillator"""
    try:
        p = prices.astype(float).dropna()
        if len(p) < k_period:
            neutral = pd.Series([50.0] * len(prices), index=prices.index)
            return neutral, neutral
        
        low_min = p.rolling(window=k_period, min_periods=k_period).min()
        high_max = p.rolling(window=k_period, min_periods=k_period).max()
        
        denom = (high_max - low_min)
        denom = denom.replace(0, np.nan)
        
        k = 100 * (p - low_min) / denom
        k = k.fillna(50.0)
        d = k.rolling(window=d_period, min_periods=1).mean()
        
        return k, d
    except Exception as e:
        logger.error(f"Stochastic calculation failed: {e}")
        neutral = pd.Series([50.0] * len(prices), index=prices.index)
        return neutral, neutral

# =======================
# CANDLESTICK PATTERNS
# =======================
def detect_candlesticks(series):
    """Enhanced candlestick pattern detection"""
    if series is None or len(series.dropna()) < 8:
        return 0.0
    
    try:
        s = series.dropna().astype(float)
        n = min(32, len(s))
        tail = s.tail(n).values
        
        # Create candlestick bars
        chunks = 8
        step = max(1, n // chunks)
        bars = []
        
        for i in range(chunks):
            start_idx = i * step
            end_idx = min((i + 1) * step, n) if i < chunks - 1 else n
            seg = tail[start_idx:end_idx]
            
            if len(seg) == 0:
                continue
                
            o = float(seg[0])
            c = float(seg[-1])
            h = float(np.max(seg))
            l = float(np.min(seg))
            bars.append((o, h, l, c))
        
        if len(bars) < 3:
            return 0.0
        
        return _analyze_candlestick_patterns(bars)
    except Exception as e:
        logger.error(f"Candlestick detection failed: {e}")
        return 0.0

def _analyze_candlestick_patterns(bars):
    """Analyze candlestick patterns for bullish/bearish signals"""
    def body(o, c): return abs(c - o)
    def upper(o, h, c): return h - max(o, c)
    def lower(o, l, c): return min(o, c) - l
    def is_bull(o, c): return c > o
    def is_bear(o, c): return c < o
    
    score = 0.0
    
    for i in range(1, len(bars) - 1):
        o1, h1, l1, c1 = bars[i-1]
        o2, h2, l2, c2 = bars[i]
        o3, h3, l3, c3 = bars[i+1] if i+1 < len(bars) else bars[i]
        
        # Calculate pattern characteristics
        b2 = body(o2, c2)
        range2 = max(h2 - l2, 1e-9)
        long_lower = (lower(o2, l2, c2) / range2) > 0.5
        long_upper = (upper(o2, h2, c2) / range2) > 0.5
        small_body = (b2 / range2) < 0.2
        
        # Single candle patterns
        if long_lower and not long_upper and is_bull(o2, c2):  # Hammer
            score += 0.15
        if long_upper and not long_lower and is_bear(o2, c2):  # Shooting star
            score -= 0.15
        if small_body:  # Doji-like
            score += 0.02 if is_bear(o1, c1) else -0.02
        
        # Multi-candle patterns
        if is_bear(o1, c1) and is_bull(o2, c2) and o2 < c1 and c2 > o1:  # Bullish engulfing
            score += 0.25
        if is_bull(o1, c1) and is_bear(o2, c2) and o2 > c1 and c2 < o1:  # Bearish engulfing
            score -= 0.25
    
    return max(-1.0, min(1.0, score))

# =======================
# PREDICTION MODELS
# =======================
def _tanh_clamp(x, limit):
    """Apply tanh clamping to prevent extreme predictions"""
    return float(math.tanh(x / max(1e-9, limit)) * limit)

def _pct_change(a, b):
    """Calculate percentage change safely"""
    if b == 0 or b is None or a is None:
        return 0.0
    return (a / b) - 1.0

def predict_5m_price(current_price, minute_series, signal_score):
    """Enhanced 5-minute price prediction"""
    w = PREDICT_WEIGHTS
    
    if minute_series is None or len(minute_series.dropna()) < 8:
        base_ret = signal_score * w["d_signal"]
        pred_ret = _tanh_clamp(base_ret, w["max_5m"])
        return current_price * (1 + pred_ret), pred_ret
    
    try:
        m = minute_series.dropna().astype(float)
        r = m.pct_change().dropna()
        
        # Recent momentum indicators
        r1 = float(r.iloc[-1]) if len(r) >= 1 else 0.0
        r5 = float(r.tail(5).sum()) if len(r) >= 5 else float(r.sum())
        
        # MACD from minute data
        ema_fast = m.ewm(span=5, adjust=False).mean()
        ema_slow = m.ewm(span=20, adjust=False).mean()
        macd_m = float((ema_fast.iloc[-1] - ema_slow.iloc[-1]) / max(1e-9, current_price))
        
        # RSI normalization
        rsi_m_series = calculate_rsi(m, period=14).dropna()
        rsi_m = float(rsi_m_series.iloc[-1]) if not rsi_m_series.empty else 50.0
        rsi_m_z = (rsi_m - 50.0) / 50.0
        
        # Volatility adjustment
        rv = float(r.tail(30).std()) if len(r) >= 10 else float(r.std() or 0.0)
        vol_shrink = 1.0 / (1.0 + 5.0 * max(0.0, rv))
        
        # Combine signals
        raw = (
            w["m_r5"] * r5 +
            w["m_macd"] * macd_m +
            w["m_rsi"] * rsi_m_z +
            w["m_r1"] * r1 +
            (signal_score * w["d_signal"])
        ) * vol_shrink
        
        pred_ret = _tanh_clamp(raw, w["max_5m"])
        return current_price * (1 + pred_ret), pred_ret
        
    except Exception as e:
        logger.error(f"5m prediction failed: {e}")
        base_ret = signal_score * w["d_signal"]
        pred_ret = _tanh_clamp(base_ret, w["max_5m"])
        return current_price * (1 + pred_ret), pred_ret

def predict_weekly_price(current_price, daily_series, signal_score):
    """Enhanced weekly price prediction using regression and momentum"""
    w = PREDICT_WEIGHTS
    d = daily_series.dropna().astype(float)
    
    if len(d) < 8:
        naive = 0.0
        if len(d) >= 2:
            naive = _pct_change(float(d.iloc[-1]), float(d.iloc[max(0, len(d) - 8)]))
        raw = w["weekly_alpha"] * naive + w["weekly_beta"] * (signal_score / 5.0) * 0.05
        pred_ret = _tanh_clamp(raw, w["weekly_cap"])
        return current_price * (1 + pred_ret), pred_ret
    
    try:
        # Regression-based trend analysis
        last_n = d.tail(12)
        x = np.arange(len(last_n), dtype=float)
        y = np.log(last_n.values + 1e-10)  # Avoid log(0)
        
        slope, intercept = np.polyfit(x, y, 1)
        reg_now = math.exp(intercept + slope * (len(last_n) - 1))
        reg_7d = math.exp(intercept + slope * (len(last_n) - 1 + 7))
        reg_ret = _pct_change(reg_7d, reg_now)
        
        # Momentum component
        naive_ref = float(d.iloc[max(0, len(d) - 8)])
        naive_ret = _pct_change(float(d.iloc[-1]), naive_ref)
        
        # Blend trend and momentum
        blended = w["weekly_alpha"] * reg_ret + (1.0 - w["weekly_alpha"]) * naive_ret
        blended += w["weekly_beta"] * (signal_score / 5.0) * 0.05
        
        pred_ret = _tanh_clamp(blended, w["weekly_cap"])
        return current_price * (1 + pred_ret), pred_ret
        
    except Exception as e:
        logger.error(f"Weekly prediction failed: {e}")
        pred_ret = _tanh_clamp(signal_score * 0.01, w["weekly_cap"])
        return current_price * (1 + pred_ret), pred_ret

# =======================
# STATE MANAGEMENT & EVALUATION
# =======================
def load_json(path, default):
    """Load JSON file with error handling"""
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load {path}: {e}")
    return default

def save_json(path, obj):
    """Save JSON file with error handling"""
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2)
    except Exception as e:
        logger.error(f"Failed to save {path}: {e}")

def append_eval_log(rows):
    """Append evaluation results to CSV log"""
    new = not os.path.exists(EVAL_LOG)
    try:
        with open(EVAL_LOG, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if new:
                writer.writerow([
                    "ts_evaluated", "ticker", "base_price", "pred_5m", 
                    "actual_now", "pred_dir", "actual_dir", "abs_err_pct"
                ])
            for row in rows:
                writer.writerow(row)
    except Exception as e:
        logger.error(f"Failed to append eval log: {e}")

def evaluator_autotune(pred_state, current_prices):
    """Evaluate predictions and auto-tune weights"""
    if not pred_state:
        return
    
    lr = CONFIG["LEARNING_RATE"]
    eval_rows = []
    correct = 0
    total = 0
    overshoot = 0.0
    overshoot_n = 0
    
    for ticker, rec in pred_state.items():
        base = rec.get("base_price")
        predp = rec.get("pred_5m")
        nowp = current_prices.get(ticker)
        
        if base is None or predp is None or nowp is None:
            continue
            
        pred_dir = 1 if (predp - base) > 0 else (-1 if (predp - base) < 0 else 0)
        actual_dir = 1 if (nowp - base) > 0 else (-1 if (nowp - base) < 0 else 0)
        
        if pred_dir == actual_dir:
            correct += 1
        total += 1
        
        err_pct = abs((nowp - predp) / max(1e-9, predp))
        overshoot += err_pct
        overshoot_n += 1
        
        eval_rows.append([
            datetime.now().isoformat(timespec="seconds"),
            ticker,
            round(base, 2),
            round(predp, 2),
            round(nowp, 2),
            "Up" if pred_dir > 0 else ("Down" if pred_dir < 0 else "Flat"),
            "Up" if actual_dir > 0 else ("Down" if actual_dir < 0 else "Flat"),
            round(err_pct * 100, 2)
        ])
    
    if eval_rows:
        append_eval_log(eval_rows)
    
    if total < CONFIG["EVAL_WINDOW_MIN"]:
        return
    
    acc = correct / total
    mean_overshoot = (overshoot / overshoot_n) if overshoot_n else 0.0
    
    logger.info(f"Prediction accuracy: {acc:.2%}, Mean error: {mean_overshoot:.3%}")
    
    # Auto-tune weights based on performance
    if acc < 0.55:
        PREDICT_WEIGHTS["m_r5"] = min(0.70, PREDICT_WEIGHTS["m_r5"] + lr * 0.15)
        PREDICT_WEIGHTS["m_r1"] = min(0.30, PREDICT_WEIGHTS["m_r1"] + lr * 0.10)
        PREDICT_WEIGHTS["m_macd"] = max(0.10, PREDICT_WEIGHTS["m_macd"] - lr * 0.10)
        PREDICT_WEIGHTS["d_signal"] = max(0.001, PREDICT_WEIGHTS["d_signal"] - lr * 0.0005)
        logger.info("Adjusted weights for better direction accuracy")
    
    if mean_overshoot > 0.006:
        PREDICT_WEIGHTS["max_5m"] = max(0.0025, PREDICT_WEIGHTS["max_5m"] - lr * 0.0005)
        logger.info("Reduced prediction magnitude due to high error")
    
    save_json(WEIGHTS_FILE, PREDICT_WEIGHTS)

# =======================
# EMAIL SYSTEM (macOS Compatible)
# =======================
def send_email(rows, attachment_path):
    """Send email notifications using SMTP (cross-platform)"""
    if not rows or not CONFIG["EMAIL_SETTINGS"]["EMAIL_USER"]:
        logger.warning("Email not configured or no rows to send")
        return
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = CONFIG["EMAIL_SETTINGS"]["EMAIL_USER"]
        msg['To'] = ", ".join(CONFIG["EMAILS"])
        msg['Subject'] = f"Crypto Signal Alert - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Create body
        body = "🚨 Crypto Assets with Strong Buy/Sell Signals:\n\n"
        for r in rows:
            if len(r) > 6:
                body += f"• {r[0]} - {r[6]} - 5m Prediction: ${r[10] if len(r) > 10 else 'N/A'}\n"
        
        body += f"\n📊 Analysis completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        body += f"\n📁 Full report attached: {os.path.basename(attachment_path)}"
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Add attachment
        if os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(attachment_path)}'
                )
                msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(CONFIG["EMAIL_SETTINGS"]["SMTP_SERVER"], CONFIG["EMAIL_SETTINGS"]["SMTP_PORT"])
        server.starttls()
        server.login(CONFIG["EMAIL_SETTINGS"]["EMAIL_USER"], CONFIG["EMAIL_SETTINGS"]["EMAIL_PASS"])
        
        text = msg.as_string()
        server.sendmail(CONFIG["EMAIL_SETTINGS"]["EMAIL_USER"], CONFIG["EMAILS"], text)
        server.quit()
        
        logger.info(f"Email sent successfully to {len(CONFIG['EMAILS'])} recipients")
        
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

# =======================
# FILE OPERATIONS
# =======================
def save_with_retries(df, path):
    """Save DataFrame to Excel with retry logic"""
    base, ext = os.path.splitext(path)
    
    for i in range(3):
        try:
            df.to_excel(path, index=False, engine='openpyxl')
            logger.info(f"Saved report to: {path}")
            return path
        except Exception as e:
            logger.warning(f"Save attempt {i+1} failed: {e}")
            if i < 2:
                time.sleep(1)
                path = f"{base}_retry{i+1}{ext}"
    
    # Final attempt
    try:
        df.to_excel(path, index=False, engine='openpyxl')
        return path
    except Exception as e:
        logger.error(f"All save attempts failed: {e}")
        # Fallback to CSV
        csv_path = path.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False)
        logger.info(f"Saved as CSV fallback: {csv_path}")
        return csv_path

def highlight_and_collect(ws):
    """Apply conditional formatting to Excel worksheet"""
    try:
        from openpyxl.styles import PatternFill
        
        red = PatternFill(start_color="FFC7CE", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", fill_type="solid")
        
        changed_rows = []
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if len(row) > 6:
                rec = str(row[6].value)  # Recommendation column
                if "Strong Buy" in rec:
                    for cell in row:
                        cell.fill = green
                    changed_rows.append([cell.value for cell in row])
                elif "Buy" in rec:
                    for cell in row:
                        cell.fill = yellow
                    changed_rows.append([cell.value for cell in row])
                elif "Strong Sell" in rec or "Sell" in rec:
                    for cell in row:
                        cell.fill = red
                    changed_rows.append([cell.value for cell in row])
        
        return changed_rows
        
    except Exception as e:
        logger.error(f"Failed to apply highlighting: {e}")
        return []

def load_last_state():
    """Load last recommendation state"""
    return load_json(LAST_STATE_FILE, {})

def save_last_state(state):
    """Save current recommendation state"""
    save_json(LAST_STATE_FILE, state)

def filter_changed_recommendations(df, last_state):
    """Find recommendations that have changed since last run"""
    changed = []
    new_state = last_state.copy()
    
    for _, row in df.iterrows():
        ticker = str(row["Ticker"])
        rec = str(row["Recommendation"])
        prev = last_state.get(ticker)
        
        if prev != rec and ("Buy" in rec or "Sell" in rec):
            changed.append(row.tolist())
        
        new_state[ticker] = rec
    
    return changed, new_state

# =======================
# CORE ANALYSIS FUNCTION
# =======================
def analyze_crypto(ticker, coin_id, current_map):
    """Main crypto analysis function with enhanced error handling"""
    try:
        # Get current price and volume
        current_price, volume = current_map.get(coin_id, (None, None))
        source = "CoinGecko"
        
        if current_price is None:
            # Fallback to Yahoo Finance
            yfp = yfinance_last_price(ticker)
            if yfp is not None:
                current_price = yfp
                volume = volume or 0
                source = "Yahoo Finance"
            else:
                logger.warning(f"No price data available for {ticker}")
                return ["N/A"] * 16
        
        # Fetch historical data
        daily = fetch_historical_series(coin_id, days=30)
        if daily.empty or len(daily) < 10:
            logger.warning(f"Insufficient daily data for {ticker}")
            return ["N/A"] * 16
        
        daily_clean = daily.dropna()
        
        # Get minute data for 5m predictions
        minute_series = None
        if CONFIG["USE_MINUTE_FOR_5M"]:
            minute_series = fetch_historical_series(coin_id, days=1)
        
        # Calculate technical indicators
        rsi_series = calculate_rsi(daily_clean)
        macd_hist = calculate_macd(daily_clean)
        k, d = calculate_stochastic_from_price(daily_clean)
        
        # Get latest indicator values
        rsi = float(rsi_series.dropna().iloc[-1]) if not rsi_series.dropna().empty else 50.0
        macd = float(macd_hist.dropna().iloc[-1]) if not macd_hist.dropna().empty else 0.0
        stochastic = float(k.dropna().iloc[-1]) if not k.dropna().empty else 50.0
        
        # Candlestick analysis
        candle_min = detect_candlesticks(minute_series) if minute_series is not None else 0.0
        candle_day = detect_candlesticks(daily_clean)
        
        # Calculate signal score
        signal_score = 0.0
        
        # MACD signal
        if macd > 0:
            signal_score += 1.2
        elif macd < 0:
            signal_score -= 0.5
        
        # RSI signals
        if rsi > 70:
            signal_score += 1.0  # Overbought but momentum
        elif rsi > 60:
            signal_score += 0.5
        elif rsi < 30:
            signal_score -= 1.0  # Oversold
        elif rsi < 40:
            signal_score -= 0.5
        
        # Stochastic signals
        if stochastic > 80:
            signal_score += 0.8
        elif stochastic < 20:
            signal_score -= 0.8
        
        # Add candlestick signals
        signal_score += PREDICT_WEIGHTS["candle_w_min"] * candle_min
        signal_score += PREDICT_WEIGHTS["candle_w_day"] * candle_day
        
        signal_score = round(signal_score, 2)
        
        # Generate recommendation
        if signal_score >= 3.5:
            recommendation = "Strong Buy"
        elif signal_score >= 2.0:
            recommendation = "Buy"
        elif signal_score <= -3.5:
            recommendation = "Strong Sell"
        elif signal_score <= -2.0:
            recommendation = "Sell"
        else:
            recommendation = "Hold"
        
        # Price predictions
        expected_weekly_price, weekly_ret = predict_weekly_price(current_price, daily_clean, signal_score)
        expected_5m_price, _ = predict_5m_price(current_price, minute_series, signal_score)
        
        # Calculate growth percentages
        weekly_growth = round(weekly_ret * 100.0, 2)
        monthly_growth = 0.0
        if len(daily_clean) >= 30:
            monthly_growth = round(((current_price - daily_clean.iloc[0]) / daily_clean.iloc[0]) * 100, 2)
        
        # Direction and call type
        direction = "Rise" if expected_5m_price > current_price else "Fall"
        call_type = "Call" if expected_5m_price > current_price else "Put"
        
        # Risk assessment
        volatility = daily_clean.pct_change().std() * np.sqrt(252) * 100  # Annualized volatility
        risk_level = "High" if volatility > 100 else ("Medium" if volatility > 50 else "Low")
        
        return [
            round(current_price, 2),
            round(volume or 0, 0),
            round(macd, 4),
            round(stochastic, 2),
            round(signal_score, 2),
            recommendation,
            f"{weekly_growth}%",
            f"{monthly_growth}%",
            round(expected_weekly_price, 2),
            round(expected_5m_price, 2),
            round(rsi, 2),
            source,
            call_type,
            direction,
            round(volatility, 2),
            risk_level
        ]
        
    except Exception as e:
        logger.error(f"Analysis failed for {ticker}: {e}")
        return ["ERROR"] * 16

# =======================
# MAIN EXECUTION FUNCTION
# =======================
def run_crypto_model():
    """Main execution function"""
    start_time = time.time()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    logger.info(f"🚀 Starting crypto analysis at {timestamp}")
    
    try:
        # Load previous predictions for evaluation
        last_pred_state = load_json(PRED_STATE_FILE, {})
        
        # Fetch current prices for all coins
        logger.info("📊 Fetching current market data...")
        current_map = fetch_current_prices_batched(list(CRYPTO_TICKERS.values()))
        
        # Map prices by ticker for evaluation
        current_prices_by_ticker = {}
        for ticker, coin_id in CRYPTO_TICKERS.items():
            price, _ = current_map.get(coin_id, (None, None))
            if price is None:
                price = yfinance_last_price(ticker)
            current_prices_by_ticker[ticker] = price
        
        # Run evaluator and auto-tune weights
        logger.info("🔧 Evaluating previous predictions and tuning model...")
        try:
            evaluator_autotune(last_pred_state, current_prices_by_ticker)
        except Exception as e:
            logger.error(f"Evaluator failed: {e}")
        
        # Analyze each cryptocurrency
        logger.info("🔍 Analyzing cryptocurrencies...")
        results = []
        new_pred_state = {}
        
        for ticker, coin_id in CRYPTO_TICKERS.items():
            logger.info(f"  → Analyzing {ticker}")
            
            row = analyze_crypto(ticker, coin_id, current_map)
            results.append([ticker] + row)
            
            # Save prediction state for next evaluation
            if row[0] != "N/A" and row[0] != "ERROR":
                try:
                    base_price = float(row[0])
                    pred_5m = float(row[9])  # Expected Price (5m)
                    new_pred_state[ticker] = {
                        "base_price": base_price,
                        "pred_5m": pred_5m,
                        "ts": datetime.now().isoformat(timespec="seconds")
                    }
                except (ValueError, IndexError):
                    pass
            
            time.sleep(0.2)  # Rate limiting
        
        # Create DataFrame
        columns = [
            "Ticker", "Price ($)", "Volume (24h)", "MACD", "Stochastic (%)", 
            "Signal Score", "Recommendation", "Weekly Growth (%)", "Monthly Growth (%)",
            "Expected Price (Weekly)", "Expected Price (5m)", "RSI", "Source",
            "Options Type", "Direction", "Volatility (%)", "Risk Level"
        ]
        
        df = pd.DataFrame(results, columns=columns)
        
        # Save to Excel
        timestamp_file = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(CONFIG["OUTPUT_FOLDER"], f"CryptoBroker_Analysis_{timestamp_file}.xlsx")
        output_path = save_with_retries(df, output_path)
        
        # Apply formatting and collect significant signals
        try:
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            changed_rows_visual = highlight_and_collect(ws)
            wb.save(output_path)
        except Exception as e:
            logger.error(f"Failed to apply Excel formatting: {e}")
            changed_rows_visual = []
        
        # Check for recommendation changes and send alerts
        last_state = load_last_state()
        changed_df_rows, new_state = filter_changed_recommendations(df, last_state)
        save_last_state(new_state)
        
        # Send email if there are significant changes
        if changed_df_rows and CONFIG["EMAIL_SETTINGS"]["EMAIL_USER"]:
            logger.info("📧 Sending email alert for significant changes...")
            send_email(changed_rows_visual, output_path)
        
        # Save prediction state for next run
        save_json(PRED_STATE_FILE, new_pred_state)
        
        # Summary statistics
        duration = round(time.time() - start_time, 2)
        strong_signals = sum("Strong" in str(rec) for rec in df["Recommendation"])
        buy_signals = sum("Buy" in str(rec) for rec in df["Recommendation"])
        sell_signals = sum("Sell" in str(rec) for rec in df["Recommendation"])
        
        logger.info(f"✅ Analysis complete!")
        logger.info(f"   📁 Report saved: {output_path}")
        logger.info(f"   📊 Assets analyzed: {len(df)}")
        logger.info(f"   🔥 Strong signals: {strong_signals}")
        logger.info(f"   📈 Buy signals: {buy_signals}")
        logger.info(f"   📉 Sell signals: {sell_signals}")
        logger.info(f"   ⏱️  Duration: {duration}s")
        
        return output_path
        
    except Exception as e:
        logger.error(f"❌ Critical error in main execution: {e}")
        raise

def sleep_to_next_five_minutes():
    """Sleep until the next 5-minute boundary"""
    now = datetime.now()
    minute = (now.minute // 5 + 1) * 5
    next_tick = now.replace(minute=0, second=0, microsecond=0) + timedelta(minutes=minute)
    
    if minute >= 60:
        next_tick = next_tick + timedelta(hours=1)
        next_tick = next_tick.replace(minute=minute - 60)
    else:
        next_tick = next_tick.replace(minute=minute)
    
    delta = (next_tick - now).total_seconds()
    delta = max(1, min(int(delta), 300))
    
    logger.info(f"⏳ Sleeping {delta} seconds until next 5-minute boundary ({next_tick.strftime('%H:%M')})")
    time.sleep(delta)

# =======================
# MAIN EXECUTION
# =======================
if __name__ == "__main__":
    logger.info("🎯 CryptoBroker Auto Trader - macOS Edition")
    logger.info(f"📁 Output folder: {CONFIG['OUTPUT_FOLDER']}")
    
    # Load email configuration
    email_configured = load_email_config()
    if not email_configured:
        logger.warning("⚠️  Email not configured - notifications disabled")
    
    # Main execution loop
    while True:
        try:
            run_crypto_model()
        except KeyboardInterrupt:
            logger.info("👋 Shutting down CryptoBroker...")
            break
        except Exception as e:
            logger.error(f"❌ Execution failed: {e}")
            logger.info("🔄 Continuing with next cycle...")
        
        # Sleep until next execution
        if CONFIG["ALIGN_TO_WALLCLOCK"]:
            sleep_to_next_five_minutes()
        else:
            logger.info("⏳ Sleeping 5 minutes...")
            time.sleep(300)
